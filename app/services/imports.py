"""Service d'import Excel des données financières (BIZ-36).

Parse un classeur Excel multi-colonnes, en extrait les hypothèses financières,
les persiste (comme une saisie manuelle) et conserve le fichier d'origine.

Principe : l'analyse est purement déterministe côté backend. Aucune donnée n'est
inventée ; si une valeur attendue est absente, l'import échoue avec un message
explicite plutôt que de produire une hypothèse erronée.
"""

from __future__ import annotations

import io
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass

from openpyxl import load_workbook
from sqlalchemy.orm import Session

from app.models import FinancialImport, FinancialStatement
from app.schemas.financial import FinancialAssumptionCreate
from app.services.financials import save_financials
from app.services.projects import get_project

#: Taille maximale acceptée pour le fichier importé (2 Mio).
MAX_IMPORT_BYTES = 2 * 1024 * 1024

#: Extensions de fichier autorisées.
_ALLOWED_EXTENSIONS = (".xlsx", ".xlsm")

#: Types MIME autorisés (Excel moderne).
_ALLOWED_CONTENT_TYPES = frozenset(
    {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel.sheet.macroEnabled.12",
        "application/octet-stream",
        "",
    }
)


class ExcelImportError(Exception):
    """Levée lorsqu'un fichier Excel est invalide ou inexploitable."""


def _normalize(label: object) -> str:
    """Normalise un libellé : minuscules, sans accents ni espaces superflus.

    Args:
        label: Valeur de cellule à normaliser.

    Returns:
        Le libellé normalisé (chaîne vide si la cellule n'est pas textuelle).
    """
    if not isinstance(label, str):
        return ""
    decomposed = unicodedata.normalize("NFKD", label)
    without_accents = "".join(c for c in decomposed if not unicodedata.combining(c))
    return without_accents.strip().lower()


def _numbers(cells: Iterable[object]) -> list[float]:
    """Extrait les valeurs numériques d'une suite de cellules.

    Args:
        cells: Cellules d'une ligne (hors libellé).

    Returns:
        La liste des nombres trouvés, dans l'ordre.
    """
    values: list[float] = []
    for cell in cells:
        if isinstance(cell, bool):
            continue
        if isinstance(cell, (int, float)):
            values.append(float(cell))
    return values


def _mean(values: list[float]) -> float:
    """Retourne la moyenne d'une liste de nombres (0.0 si vide)."""
    return sum(values) / len(values) if values else 0.0


def parse_financials_xlsx(content: bytes) -> FinancialAssumptionCreate:
    """Extrait les hypothèses financières d'un classeur Excel.

    Le format attendu est une feuille « libellé puis valeurs » : la première
    cellule de chaque ligne porte un libellé (ex. « Revenus annuels ») et les
    cellules suivantes contiennent une ou plusieurs valeurs annuelles. Pour les
    postes pluriannuels (revenus, coûts), la moyenne des colonnes est retenue.

    Args:
        content: Contenu binaire du fichier Excel.

    Returns:
        Les hypothèses financières extraites et validées.

    Raises:
        ExcelImportError: Si le fichier est illisible ou s'il manque une donnée.
    """
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lève divers types selon la corruption.
        raise ExcelImportError("Fichier Excel illisible ou corrompu.") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise ExcelImportError("Le classeur ne contient aucune feuille.")

        investissement: float | None = None
        revenus: float | None = None
        couts: float | None = None
        delai: float | None = None

        for row in sheet.iter_rows(values_only=True):
            if not row:
                continue
            label = _normalize(row[0])
            if not label:
                continue
            nums = _numbers(row[1:])
            if not nums:
                continue
            if investissement is None and "investissement" in label:
                investissement = nums[0]
            elif revenus is None and "revenu" in label:
                revenus = _mean(nums)
            elif couts is None and ("cout" in label or "charge" in label):
                couts = _mean(nums)
            elif delai is None and ("delai" in label or "rentab" in label or "retour" in label):
                delai = nums[0]
    finally:
        workbook.close()

    missing = [
        name
        for name, value in (
            ("investissement initial", investissement),
            ("revenus annuels", revenus),
            ("coûts annuels", couts),
            ("délai de rentabilité", delai),
        )
        if value is None
    ]
    if missing:
        raise ExcelImportError(
            "Données financières manquantes dans le fichier : " + ", ".join(missing) + "."
        )

    try:
        return FinancialAssumptionCreate(
            investissement_initial=float(investissement),  # type: ignore[arg-type]
            revenus_annuels=round(float(revenus), 2),  # type: ignore[arg-type]
            couts_annuels=round(float(couts), 2),  # type: ignore[arg-type]
            delai_rentabilite_mois=int(round(float(delai))),  # type: ignore[arg-type]
        )
    except ValueError as exc:
        raise ExcelImportError("Valeurs financières invalides dans le fichier.") from exc


def _validate_upload(filename: str, content: bytes) -> None:
    """Valide l'extension et la taille du fichier téléversé.

    Args:
        filename: Nom du fichier d'origine.
        content: Contenu binaire.

    Raises:
        ExcelImportError: Si l'extension n'est pas autorisée ou le fichier vide.
        FileTooLargeError: Si le fichier dépasse la taille maximale.
    """
    name = filename.lower()
    if not name.endswith(_ALLOWED_EXTENSIONS):
        raise ExcelImportError(
            "Format non supporté : seuls les fichiers .xlsx/.xlsm sont acceptés."
        )
    if not content:
        raise ExcelImportError("Le fichier importé est vide.")
    if len(content) > MAX_IMPORT_BYTES:
        raise FileTooLargeError(len(content))


class FileTooLargeError(Exception):
    """Levée lorsqu'un fichier importé dépasse la taille maximale autorisée."""


def import_financials(
    db: Session,
    project_id: int,
    *,
    filename: str,
    content_type: str,
    content: bytes,
) -> tuple[FinancialImport, FinancialAssumptionCreate]:
    """Importe un fichier Excel : valide, parse, persiste et conserve le fichier.

    Args:
        db: Session de base de données.
        project_id: Identifiant du projet rattaché.
        filename: Nom du fichier d'origine.
        content_type: Type MIME déclaré à l'upload.
        content: Contenu binaire du fichier.

    Returns:
        Le couple ``(import persisté, hypothèses extraites)``.

    Raises:
        ProjectNotFoundError: Si le projet n'existe pas.
        ExcelImportError: Si le fichier est invalide ou inexploitable.
        FileTooLargeError: Si le fichier dépasse la taille maximale.
    """
    project = get_project(db, project_id)
    _validate_upload(filename, content)

    parsed = parse_financials_xlsx(content)
    save_financials(db, project_id, parsed)

    record = project.financial_import
    if record is None:
        record = FinancialImport(project_id=project_id)
        db.add(record)
    record.filename = filename[:255]
    record.content_type = (content_type or "application/octet-stream")[:100]
    record.size_bytes = len(content)
    record.content = content

    db.commit()
    db.refresh(record)
    return record, parsed


def get_import(db: Session, project_id: int) -> FinancialImport:
    """Retourne le dernier fichier Excel importé pour un projet.

    Args:
        db: Session de base de données.
        project_id: Identifiant du projet rattaché.

    Returns:
        Le fichier importé persisté.

    Raises:
        ProjectNotFoundError: Si le projet n'existe pas.
        ImportNotFoundError: Si aucun fichier n'a été importé.
    """
    project = get_project(db, project_id)
    if project.financial_import is None:
        raise ImportNotFoundError(project_id)
    return project.financial_import


class ImportNotFoundError(Exception):
    """Levée lorsqu'aucun fichier Excel n'a été importé pour le projet."""


# ---------------------------------------------------------------------------
# Import détaillé multi-colonnes (BIZ-32) : temps en lignes, catégories en
# colonnes (dépenses, recettes, agrégats), conformément à docs/Projet2.docx.
# ---------------------------------------------------------------------------

#: Correspondance ``(fragment de libellé normalisé, groupe, clé canonique)``.
#: L'ordre privilégie les fragments les plus spécifiques ; la première
#: correspondance trouvée pour un en-tête de colonne l'emporte.
_COLUMN_MAP: tuple[tuple[str, str, str], ...] = (
    ("total depense", "depenses", "total_depenses"),
    ("autres depense", "depenses", "autres_depenses"),
    ("salaire", "depenses", "salaires"),
    ("materiel", "depenses", "achat_materiel"),
    ("logiciel", "depenses", "achat_logiciel"),
    ("fiscal", "depenses", "charges_fiscales"),
    ("administ", "depenses", "frais_administratifs"),
    ("bancaire", "depenses", "frais_bancaires"),
    ("divers", "depenses", "achats_divers"),
    ("nombre de client", "recettes", "nombre_clients"),
    ("nombre client", "recettes", "nombre_clients"),
    ("nb client", "recettes", "nombre_clients"),
    ("service 1", "recettes", "recette_produit_1"),
    ("produit 1", "recettes", "recette_produit_1"),
    ("service 2", "recettes", "recette_produit_2"),
    ("produit 2", "recettes", "recette_produit_2"),
    ("service 3", "recettes", "recette_produit_3"),
    ("produit 3", "recettes", "recette_produit_3"),
    ("service 4", "recettes", "recette_produit_4"),
    ("produit 4", "recettes", "recette_produit_4"),
    ("chiffre", "recettes", "chiffre_affaires"),
    ("marge", "agregats", "marge_brute"),
    ("ebitda", "agregats", "ebitda"),
    ("exploitation", "agregats", "resultat_exploitation"),
    ("ebe", "agregats", "ebe"),
)

#: Postes de dépenses agrégés pour reconstituer le total si absent du fichier.
_EXPENSE_KEYS = (
    "salaires",
    "achat_materiel",
    "achat_logiciel",
    "charges_fiscales",
    "frais_administratifs",
    "frais_bancaires",
    "achats_divers",
    "autres_depenses",
)

#: Postes de recettes produits pour reconstituer le CA si absent du fichier.
_REVENUE_KEYS = (
    "recette_produit_1",
    "recette_produit_2",
    "recette_produit_3",
    "recette_produit_4",
)

#: Nombre de mois représenté par une période, selon la granularité détectée.
_MONTHS_PER_PERIOD: dict[str, float] = {
    "semaine": 12.0 / 52.0,
    "mois": 1.0,
    "annee": 12.0,
}


@dataclass(frozen=True)
class ParsedStatement:
    """Tableau financier détaillé extrait d'un classeur Excel (BIZ-32).

    Attributes:
        period_unit: Granularité temporelle détectée.
        periods: Libellés des périodes, dans l'ordre.
        depenses: Séries de dépenses par poste canonique.
        recettes: Séries de recettes par poste canonique.
        agregats: Séries d'agrégats par poste canonique.
    """

    period_unit: str
    periods: list[str]
    depenses: dict[str, list[float]]
    recettes: dict[str, list[float]]
    agregats: dict[str, list[float]]


def _cell_number(cell: object) -> float:
    """Convertit une cellule en nombre (0.0 si non numérique)."""
    if isinstance(cell, bool):
        return 0.0
    if isinstance(cell, (int, float)):
        return float(cell)
    return 0.0


def _detect_period_unit(header_label: object, period_labels: Iterable[str]) -> str:
    """Détecte la granularité temporelle à partir des libellés.

    Args:
        header_label: Libellé de la cellule d'en-tête de la colonne temporelle.
        period_labels: Libellés des périodes (lignes).

    Returns:
        ``"semaine"``, ``"annee"`` ou ``"mois"`` (valeur par défaut).
    """
    blob = " ".join([_normalize(header_label), *(_normalize(p) for p in period_labels)])
    if "semaine" in blob:
        return "semaine"
    if "annee" in blob:
        return "annee"
    return "mois"


def _elementwise_sum(series: Iterable[list[float]], length: int) -> list[float]:
    """Additionne terme à terme plusieurs séries de même longueur.

    Args:
        series: Séries à additionner.
        length: Nombre de périodes attendu.

    Returns:
        La série somme (zéros si aucune série).
    """
    totals = [0.0] * length
    for serie in series:
        for index in range(length):
            totals[index] += serie[index]
    return totals


def parse_financial_statement(content: bytes) -> ParsedStatement:
    """Extrait un tableau financier détaillé d'un classeur Excel (BIZ-32).

    Format attendu (cf. docs/Projet2.docx) : le temps en lignes (semaines, mois
    ou années) et les catégories en colonnes — dépenses, recettes et agrégats.
    La première ligne porte les libellés de colonnes ; la première cellule de
    chaque ligne suivante porte le libellé de la période.

    Args:
        content: Contenu binaire du fichier Excel.

    Returns:
        Le tableau financier structuré.

    Raises:
        ExcelImportError: Si le fichier est illisible ou si les colonnes
            attendues sont absentes.
    """
    try:
        workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    except Exception as exc:  # openpyxl lève divers types selon la corruption.
        raise ExcelImportError("Fichier Excel illisible ou corrompu.") from exc

    try:
        sheet = workbook.active
        if sheet is None:
            raise ExcelImportError("Le classeur ne contient aucune feuille.")
        rows = [
            row
            for row in sheet.iter_rows(values_only=True)
            if row and any(cell is not None for cell in row)
        ]
    finally:
        workbook.close()

    if not rows:
        raise ExcelImportError("Le fichier ne contient aucune donnée.")

    header = rows[0]
    column_map: dict[int, tuple[str, str]] = {}
    for col_index, label in enumerate(header[1:], start=1):
        normalized = _normalize(label)
        if not normalized:
            continue
        for fragment, group, key in _COLUMN_MAP:
            if fragment in normalized and key not in {k for _, k in column_map.values()}:
                column_map[col_index] = (group, key)
                break

    if not column_map:
        raise ExcelImportError(
            "Aucune colonne de catégorie reconnue (dépenses, recettes ou agrégats)."
        )

    periods: list[str] = []
    series: dict[tuple[str, str], list[float]] = {gk: [] for gk in column_map.values()}

    for row in rows[1:]:
        label = row[0] if row else None
        if label is None:
            continue
        periods.append(str(label).strip())
        for col_index, group_key in column_map.items():
            cell = row[col_index] if col_index < len(row) else None
            series[group_key].append(_cell_number(cell))

    if not periods:
        raise ExcelImportError("Aucune période (ligne de données) trouvée dans le fichier.")

    depenses = {key: values for (group, key), values in series.items() if group == "depenses"}
    recettes = {key: values for (group, key), values in series.items() if group == "recettes"}
    agregats = {key: values for (group, key), values in series.items() if group == "agregats"}

    has_revenue = "chiffre_affaires" in recettes or any(k in recettes for k in _REVENUE_KEYS)
    has_expense = "total_depenses" in depenses or any(k in depenses for k in _EXPENSE_KEYS)
    if not has_revenue:
        raise ExcelImportError(
            "Colonne de recettes manquante (chiffre d'affaires ou recette produit/service)."
        )
    if not has_expense:
        raise ExcelImportError("Colonne de dépenses manquante.")

    unit = _detect_period_unit(header[0], periods)
    return ParsedStatement(
        period_unit=unit,
        periods=periods,
        depenses=depenses,
        recettes=recettes,
        agregats=agregats,
    )


def derive_assumptions(statement: ParsedStatement) -> FinancialAssumptionCreate:
    """Dérive les hypothèses financières scalaires d'un tableau détaillé (BIZ-32).

    Les séries sont agrégées de façon déterministe : revenus et coûts sont
    annualisés (ramenés à 12 mois selon la granularité), l'investissement
    initial correspond au besoin de financement maximal (creux de trésorerie
    cumulée), et le délai de rentabilité au premier moment où la trésorerie
    cumulée redevient positive.

    Args:
        statement: Tableau financier détaillé extrait du fichier.

    Returns:
        Les hypothèses financières dérivées et validées.
    """
    count = len(statement.periods)
    months_per_period = _MONTHS_PER_PERIOD.get(statement.period_unit, 1.0)

    ca = statement.recettes.get("chiffre_affaires")
    if ca is None:
        ca = _elementwise_sum(
            (statement.recettes[k] for k in _REVENUE_KEYS if k in statement.recettes),
            count,
        )
    depenses = statement.depenses.get("total_depenses")
    if depenses is None:
        depenses = _elementwise_sum(
            (statement.depenses[k] for k in _EXPENSE_KEYS if k in statement.depenses),
            count,
        )

    total_ca = sum(ca)
    total_dep = sum(depenses)
    months_total = max(count * months_per_period, months_per_period)
    annual_factor = 12.0 / months_total

    revenus_annuels = max(round(total_ca * annual_factor, 2), 0.0)
    couts_annuels = max(round(total_dep * annual_factor, 2), 0.0)

    cumulative = 0.0
    min_cumulative = 0.0
    payback_period: int | None = None
    for index in range(count):
        cumulative += ca[index] - depenses[index]
        min_cumulative = min(min_cumulative, cumulative)
        if payback_period is None and cumulative >= 0:
            payback_period = index + 1

    investissement_initial = round(abs(min_cumulative), 2)
    if investissement_initial == 0.0 and depenses:
        investissement_initial = round(depenses[0], 2)

    if payback_period is None:
        delai_mois = 600
    else:
        delai_mois = max(1, min(600, round(payback_period * months_per_period)))

    return FinancialAssumptionCreate(
        investissement_initial=investissement_initial,
        revenus_annuels=revenus_annuels,
        couts_annuels=couts_annuels,
        delai_rentabilite_mois=delai_mois,
    )


def import_financial_statement(
    db: Session,
    project_id: int,
    *,
    filename: str,
    content_type: str,
    content: bytes,
) -> tuple[FinancialStatement, FinancialAssumptionCreate]:
    """Importe un tableau financier détaillé Excel (BIZ-32).

    Valide le fichier, parse le tableau multi-colonnes, en dérive les
    hypothèses financières (persistées comme une saisie manuelle), conserve le
    fichier d'origine et le tableau structuré.

    Args:
        db: Session de base de données.
        project_id: Identifiant du projet rattaché.
        filename: Nom du fichier d'origine.
        content_type: Type MIME déclaré à l'upload.
        content: Contenu binaire du fichier.

    Returns:
        Le couple ``(tableau persisté, hypothèses dérivées)``.

    Raises:
        ProjectNotFoundError: Si le projet n'existe pas.
        ExcelImportError: Si le fichier est invalide ou inexploitable.
        FileTooLargeError: Si le fichier dépasse la taille maximale.
    """
    project = get_project(db, project_id)
    _validate_upload(filename, content)

    parsed = parse_financial_statement(content)
    derived = derive_assumptions(parsed)
    save_financials(db, project_id, derived)

    record = project.financial_import
    if record is None:
        record = FinancialImport(project_id=project_id)
        db.add(record)
    record.filename = filename[:255]
    record.content_type = (content_type or "application/octet-stream")[:100]
    record.size_bytes = len(content)
    record.content = content

    statement = project.financial_statement
    if statement is None:
        statement = FinancialStatement(project_id=project_id)
        db.add(statement)
    statement.period_unit = parsed.period_unit
    statement.periods = parsed.periods
    statement.depenses = parsed.depenses
    statement.recettes = parsed.recettes
    statement.agregats = parsed.agregats

    db.commit()
    db.refresh(statement)
    return statement, derived


def get_financial_statement(db: Session, project_id: int) -> FinancialStatement:
    """Retourne le tableau financier détaillé importé pour un projet (BIZ-32).

    Args:
        db: Session de base de données.
        project_id: Identifiant du projet rattaché.

    Returns:
        Le tableau financier détaillé persisté.

    Raises:
        ProjectNotFoundError: Si le projet n'existe pas.
        StatementNotFoundError: Si aucun tableau détaillé n'a été importé.
    """
    project = get_project(db, project_id)
    if project.financial_statement is None:
        raise StatementNotFoundError(project_id)
    return project.financial_statement


class StatementNotFoundError(Exception):
    """Levée lorsqu'aucun tableau financier détaillé n'a été importé."""
